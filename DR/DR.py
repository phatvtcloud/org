import os
import glob
import re
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    # Configure stdout to handle Vietnamese unicode output in console
    sys.stdout.reconfigure(encoding='utf-8')
    
    pq_dir = r"C:\Users\phatvt\OneDrive - WIN\DR-WCM - 15.9.3 Operation\DR\pq"
    
    print("Đang quét thư mục:", pq_dir)
    if not os.path.exists(pq_dir):
        print(f"Lỗi: Thư mục {pq_dir} không tồn tại.")
        return
        
    # Lấy toàn bộ file parquet trong folder
    files = glob.glob(os.path.join(pq_dir, "*.parquet"))
    
    # Định dạng tên file: YYYYMMDD_XXh.parquet (VD: 20260605_09h.parquet)
    pattern = re.compile(r"^(\d{8})_(\d+)h\.parquet$")
    
    daily_files = {}
    for f in files:
        basename = os.path.basename(f)
        match = pattern.match(basename)
        if match:
            date_str, hour_str = match.groups()
            # Lọc từ ngày 20260526 đến hiện tại
            if date_str >= "20260526":
                hour = int(hour_str)
                if date_str not in daily_files or hour > daily_files[date_str]['hour']:
                    daily_files[date_str] = {
                        'file_path': f,
                        'hour': hour,
                        'basename': basename
                    }
                    
    if not daily_files:
        print("Không tìm thấy file nào thỏa mãn điều kiện từ ngày 20260526.")
        return
        
    print("\nCác file được chọn (mỗi ngày lấy 1 file mới nhất):")
    sorted_dates = sorted(daily_files.keys())
    for d in sorted_dates:
        info = daily_files[d]
        print(f"  - Ngày {d}: {info['basename']} (giờ: {info['hour']}h)")
        
    # Đọc và combine các file
    dfs = []
    for d in sorted_dates:
        file_path = daily_files[d]['file_path']
        try:
            df = pd.read_parquet(file_path, columns=['DC Name', 'DR_FLAG_SUM', 'TOTAL_LINE_SUM'])
            # Định dạng Ngày để hiển thị dạng DD/MM/YYYY
            date_formatted = f"{d[6:8]}/{d[4:6]}/{d[0:4]}"
            df['Ngày'] = date_formatted
            dfs.append(df)
        except Exception as e:
            print(f"Lỗi khi đọc file {file_path}: {e}")
            
    if not dfs:
        print("Không có dữ liệu được load.")
        return
        
    combined_df = pd.concat(dfs, ignore_index=True)
    print(f"\nĐã combine {len(dfs)} file parquet. Tổng số dòng: {len(combined_df):,}")
    
    # Group và tính tổng
    grouped = combined_df.groupby(['DC Name', 'Ngày']).agg(
        DR_FLAG_SUM=('DR_FLAG_SUM', 'sum'),
        TOTAL_LINE_SUM=('TOTAL_LINE_SUM', 'sum')
    ).reset_index()
    
    # Tính toán Measure %DR = SUM(DR_FLAG_SUM) / SUM(TOTAL_LINE_SUM)
    grouped['DR_Rate'] = grouped['DR_FLAG_SUM'] / grouped['TOTAL_LINE_SUM']
    
    # Pivot bảng dữ liệu
    pivot_df = grouped.pivot(index='DC Name', columns='Ngày', values='DR_Rate')
    
    # Sắp xếp các cột theo thứ tự thời gian
    columns_ordered = [f"{d[6:8]}/{d[4:6]}/{d[0:4]}" for d in sorted_dates]
    columns_ordered = [c for c in columns_ordered if c in pivot_df.columns]
    pivot_df = pivot_df[columns_ordered]
    
    # ------------------ THÊM SUBTOTAL THEO ROW VÀ COLUMN ------------------
    # 1. Tính tổng cộng theo hàng (Row Subtotal - cho từng DC Name)
    dc_totals = combined_df.groupby('DC Name').agg(
        DR_FLAG_SUM=('DR_FLAG_SUM', 'sum'),
        TOTAL_LINE_SUM=('TOTAL_LINE_SUM', 'sum')
    )
    dc_totals['Tổng cộng'] = dc_totals['DR_FLAG_SUM'] / dc_totals['TOTAL_LINE_SUM']
    pivot_df['Tổng cộng'] = dc_totals['Tổng cộng']
    
    # 2. Tính tổng cộng theo cột (Column Subtotal - cho từng Ngày)
    date_totals = combined_df.groupby('Ngày').agg(
        DR_FLAG_SUM=('DR_FLAG_SUM', 'sum'),
        TOTAL_LINE_SUM=('TOTAL_LINE_SUM', 'sum')
    )
    date_totals['DR_Rate'] = date_totals['DR_FLAG_SUM'] / date_totals['TOTAL_LINE_SUM']
    
    # 3. Tính Grand Total (giao điểm hàng và cột)
    grand_dr_flag = combined_df['DR_FLAG_SUM'].sum()
    grand_total_line = combined_df['TOTAL_LINE_SUM'].sum()
    grand_total_rate = grand_dr_flag / grand_total_line
    
    # 4. Gộp dòng Tổng cộng vào pivot_df
    col_totals_series = date_totals['DR_Rate']
    row_total_df = pd.DataFrame([col_totals_series])
    row_total_df.index = ['Tổng cộng']
    row_total_df.loc['Tổng cộng', 'Tổng cộng'] = grand_total_rate
    
    pivot_df = pd.concat([pivot_df, row_total_df])
    # ----------------------------------------------------------------------
    
    # Tạo phiên bản hiển thị phần trăm trên console
    pivot_formatted = pivot_df.copy()
    for col in pivot_formatted.columns:
        pivot_formatted[col] = pivot_formatted[col].apply(
            lambda x: f"{x*100:.1f}%" if pd.notnull(x) else "-"
        )
        
    print("\n" + "="*80)
    print("BẢNG PIVOT %DR (Dòng: DC Name, Cột: Ngày, Measure: SUM(DR_FLAG_SUM)/SUM(TOTAL_LINE_SUM))")
    print("="*80)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.width', 1000)
    print(pivot_formatted)
    print("="*80)
    
    # Xuất file Excel định dạng đẹp giống như hình
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_excel = os.path.join(output_dir, "DR_Pivot_Table.xlsx")
    
    try:
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='DR_Pivot')
            
            workbook = writer.book
            worksheet = writer.sheets['DR_Pivot']
            
            # Thiết lập style
            # Header: Màu xanh navy đậm, chữ trắng, in đậm
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            
            # Border mỏng màu xám
            border_thin = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Border cho dòng Tổng cộng (Accounting double line ở dưới)
            border_total_row = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='B0C4DE'),
                bottom=Side(style='double', color='1F497D')
            )
            
            # Áp dụng style cho Header
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_thin
            
            # Thiết lập định dạng dữ liệu (Phần trăm 0.0%) và Conditional Formatting màu sắc
            for row_idx in range(2, worksheet.max_row + 1):
                row_name = worksheet.cell(row=row_idx, column=1).value
                is_total_row = (row_name == "Tổng cộng")
                
                # Format cột DC Name
                cell_dc = worksheet.cell(row=row_idx, column=1)
                cell_dc.font = Font(name="Calibri", size=11, bold=True)
                cell_dc.alignment = Alignment(horizontal="left", vertical="center")
                cell_dc.border = border_total_row if is_total_row else border_thin
                
                if is_total_row:
                    cell_dc.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell_dc.font = Font(name="Calibri", size=11, bold=True, color="1F497D")
                
                # Format các cột giá trị
                for col_idx in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_name = worksheet.cell(row=1, column=col_idx).value
                    is_total_col = (col_name == "Tổng cộng")
                    
                    cell.number_format = '0.0%'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = border_total_row if is_total_row else border_thin
                    
                    # Xác định kiểu in đậm cho dòng/cột Tổng cộng
                    is_bold = is_total_row or is_total_col
                    
                    val = cell.value
                    if isinstance(val, (int, float)):
                        # Áp dụng Conditional Formatting màu sắc cho tất cả các ô (bao gồm cả dòng/cột Tổng cộng)
                        # >= 95% -> Xanh lá cây nhạt
                        # >= 90% -> Vàng/Cam nhạt
                        # < 90%  -> Đỏ nhạt
                        if val >= 0.95:
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                            cell.font = Font(name="Calibri", size=11, bold=is_bold, color="375623")
                        elif val >= 0.90:
                            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                            cell.font = Font(name="Calibri", size=11, bold=is_bold, color="7F6000")
                        else:
                            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                            cell.font = Font(name="Calibri", size=11, bold=is_bold, color="C65911")
                    else:
                        # Dự phòng cho trường hợp không có giá trị
                        if is_bold:
                            cell.font = Font(name="Calibri", size=11, bold=True, color="1F497D")
                            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                            
            # Auto-fit Column Widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = ""
                    if cell.value is not None:
                        if isinstance(cell.value, float) and cell.number_format == '0.0%':
                            val_str = f"{cell.value*100:.1f}%"
                        else:
                            val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
            # Row heights
            worksheet.row_dimensions[1].height = 25
            for r in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[r].height = 20
                
        print(f"\nĐã xuất bảng pivot đẹp và định dạng giống hình vào Excel:")
        print(f"-> {output_excel}")
    except Exception as e:
        print(f"Lỗi khi xuất file Excel: {e}")

if __name__ == "__main__":
    main()

